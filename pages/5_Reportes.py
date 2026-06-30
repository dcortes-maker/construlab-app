import streamlit as st
import sys; sys.path.insert(0, '..')
from auth import verificar_login, barra_superior
from utils import _proyecto_db, cargar_datos, cargar_reservas
from datetime import date
import io

st.set_page_config(page_title="Reportes", page_icon="📊", layout="wide")
verificar_login()
barra_superior()

st.markdown("## 📊 Reportes")
st.markdown("---")

datos    = cargar_datos(_proyecto_db())
hoy      = datos['hoy']
filas    = datos['filas']
clientes = datos['clientes']
reservas = cargar_reservas(_proyecto_db())

# ── Helpers Excel ──────────────────────────────────────────────────
def _wb_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def _header_style(ws, row, cols, fill_hex="0D5C6E"):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    fill   = PatternFill("solid", fgColor=fill_hex)
    font   = Font(bold=True, color="FFFFFF", size=10)
    border = Border(bottom=Side(style='thin', color='AAAAAA'))
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.fill   = fill
        c.font   = font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')

def _money(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = '"$"#,##0.00'
    return c

def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

# ── Generadores ────────────────────────────────────────────────────
def reporte_resumen():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen General"
    cols = ["Apto","Cliente","# Cuotas","Total Plan","Separación",
            "G. Manejo","G. Legal","Abono","Total Abonado","Atrasado","Estado"]
    _header_style(ws, 1, cols)
    sin_plan = datos.get('sin_plan', set())
    for r, (u, nom) in enumerate(clientes, 2):
        d   = datos['dash'].get(u, {})
        sp  = u in sin_plan
        sep = d.get('separacion', 0)
        gm  = d.get('gasto_manejo', 0)
        gl  = d.get('gasto_legal', 0)
        abo = d.get('abono', 0)
        tot = sep + gm + gl + abo
        atr = d.get('atrasado', 0)
        ws.cell(r, 1, u)
        ws.cell(r, 2, ("SIN PLAN — " if sp else "") + nom)
        ws.cell(r, 3, "—" if sp else d.get('n_cuotas', 0))
        _money(ws, r, 4, d.get('total_plan', 0) if not sp else 0)
        _money(ws, r, 5, sep); _money(ws, r, 6, gm); _money(ws, r, 7, gl)
        _money(ws, r, 8, abo); _money(ws, r, 9, tot)
        c_atr = _money(ws, r, 10, atr)
        if atr > 0: c_atr.font = Font(color="C00000", bold=True)
        ws.cell(r, 11, "Sin plan" if sp else ("Al día" if atr == 0 else "En mora"))
    _autowidth(ws)
    return _wb_bytes(wb)

def reporte_cuotas_mes(anio, mes, incluir_atrasadas):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cuotas {mes:02d}-{anio}"
    cols = ["Apto","Cliente","Descripción","F. Vencimiento","Monto","Estado","F. Pago"]
    _header_style(ws, 1, cols)
    filas_mes = []
    for f in filas:
        if f['monto'] <= 0: continue
        if f['fv']:
            es_mes = f['fv'].year == anio and f['fv'].month == mes
            es_atr = (f['fv'].year, f['fv'].month) < (hoy.year, hoy.month) and not f['fp']
            if es_mes or (incluir_atrasadas and es_atr):
                filas_mes.append(f)
    filas_mes.sort(key=lambda x: (x['unidad'], x['fv'] or date(2099,1,1)))
    for r, f in enumerate(filas_mes, 2):
        ws.cell(r, 1, f['unidad']); ws.cell(r, 2, f['nombre']); ws.cell(r, 3, f['desc'])
        ws.cell(r, 4, f['fv'].strftime('%d/%m/%Y') if f['fv'] else '—')
        _money(ws, r, 5, f['monto'])
        if f['fp']:
            ws.cell(r, 6, 'Pagado').font = Font(color="1B7A34")
            ws.cell(r, 7, f['fp'].strftime('%d/%m/%Y'))
        elif f['fv'] and (f['fv'].year, f['fv'].month) < (hoy.year, hoy.month):
            ws.cell(r, 6, 'Atrasado').font = Font(color="C00000", bold=True)
            ws.cell(r, 7, '—')
        else:
            ws.cell(r, 6, 'Pendiente'); ws.cell(r, 7, '—')
    _autowidth(ws)
    return _wb_bytes(wb)

def reporte_historial():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Pagos"
    cols = ["Apto","Cliente","Descripción","F. Vencimiento","Monto","F. Pago"]
    _header_style(ws, 1, cols)
    pagados = sorted([f for f in filas if f['fp'] and f['monto'] > 0], key=lambda x: x['fp'])
    for r, f in enumerate(pagados, 2):
        ws.cell(r, 1, f['unidad']); ws.cell(r, 2, f['nombre']); ws.cell(r, 3, f['desc'])
        ws.cell(r, 4, f['fv'].strftime('%d/%m/%Y') if f['fv'] else '—')
        _money(ws, r, 5, f['monto'])
        ws.cell(r, 6, f['fp'].strftime('%d/%m/%Y'))
    tr = len(pagados) + 2
    ws.cell(tr, 2, "TOTAL").font = Font(bold=True)
    _money(ws, tr, 5, sum(f['monto'] for f in pagados)).font = Font(bold=True)
    _autowidth(ws)
    return _wb_bytes(wb)

def reporte_mora():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes en Mora"
    cols = ["Apto","Cliente","# Cuotas Atrasadas","Total en Mora","Cuota más antigua"]
    _header_style(ws, 1, cols, fill_hex="8B0000")
    mora = {}
    for f in filas:
        if not f['fp'] and f['fv'] and f['monto'] > 0:
            if (f['fv'].year, f['fv'].month) < (hoy.year, hoy.month):
                u = f['unidad']
                if u not in mora:
                    mora[u] = {'nombre': f['nombre'], 'count': 0, 'total': 0, 'mindate': f['fv']}
                mora[u]['count'] += 1
                mora[u]['total'] += f['monto']
                if f['fv'] < mora[u]['mindate']:
                    mora[u]['mindate'] = f['fv']
    for r, (u, d) in enumerate(sorted(mora.items()), 2):
        ws.cell(r, 1, u); ws.cell(r, 2, d['nombre']); ws.cell(r, 3, d['count'])
        _money(ws, r, 4, d['total']).font = Font(color="C00000", bold=True)
        ws.cell(r, 5, d['mindate'].strftime('%d/%m/%Y'))
    _autowidth(ws)
    return _wb_bytes(wb)

def reporte_plan_por_cliente():
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    wb.remove(wb.active)
    cols = ["#","Descripción","F. Vencimiento","Monto","Estado","F. Pago"]
    for u, nom in clientes:
        ws = wb.create_sheet(title=u.replace('/', '-')[:31])
        ws.cell(1, 1, f"{nom} — Apto {u}").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        _header_style(ws, 2, cols)
        cli_filas = sorted(
            [f for f in filas if f['unidad'] == u and f['monto'] > 0 and 'bonif' not in f['desc'].lower()],
            key=lambda x: x['fv'] or date(2099,1,1)
        )
        total = 0
        for r, f in enumerate(cli_filas, 3):
            ws.cell(r, 1, r-2); ws.cell(r, 2, f['desc'])
            ws.cell(r, 3, f['fv'].strftime('%d/%m/%Y') if f['fv'] else '—')
            _money(ws, r, 4, f['monto']); total += f['monto']
            if f['fp']:
                ws.cell(r, 5, 'Pagado').font = Font(color="1B7A34")
                ws.cell(r, 6, f['fp'].strftime('%d/%m/%Y'))
            elif f['fv'] and (f['fv'].year, f['fv'].month) < (hoy.year, hoy.month):
                ws.cell(r, 5, 'Atrasado').font = Font(color="C00000", bold=True)
                ws.cell(r, 6, '—')
            else:
                ws.cell(r, 5, 'Pendiente'); ws.cell(r, 6, '—')
        tr = len(cli_filas) + 3
        ws.cell(tr, 2, "TOTAL").font = Font(bold=True)
        _money(ws, tr, 4, total).font = Font(bold=True)
        _autowidth(ws)
    return _wb_bytes(wb)

def reporte_reservas():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservas sin CPP"
    _header_style(ws, 1, ["Lote / Apto","Cliente","Monto ($)","Fecha","Notas"], fill_hex="7C5200")
    for r, res in enumerate(reservas, 2):
        ws.cell(r, 1, res['unidad']); ws.cell(r, 2, res['nombre'])
        _money(ws, r, 3, res['monto'])
        ws.cell(r, 4, res.get('fecha','')); ws.cell(r, 5, res.get('notas',''))
    _autowidth(ws)
    return _wb_bytes(wb)

# ── UI ──────────────────────────────────────────────────────────────
REPORTES = {
    "📊 Resumen General":          "Una fila por cliente con cuotas, totales abonados, atrasado y estado.",
    "📅 Cuotas del Mes":           "Cuotas que vencen en el mes seleccionado, con opción de incluir atrasadas.",
    "💳 Historial de Pagos":       "Todos los pagos registrados ordenados por fecha. Libro de ingresos.",
    "⚠️ Clientes en Mora":         "Solo clientes con cuotas vencidas: número de cuotas, monto total y cuota más antigua.",
    "📋 Plan de Pagos por Cliente": "Excel multi-hoja, una hoja por cliente con su plan completo.",
    "🔖 Clientes en Reserva":      "Lista de clientes con reserva pero sin CPP.",
}

sel = st.selectbox("Selecciona el reporte", list(REPORTES.keys()), key="sel_reporte",
                   label_visibility="collapsed",
                   placeholder="— Elige un reporte —")

if sel:
    st.markdown(f"<div style='color:#64748b;font-size:.88rem;margin:6px 0 20px;'>{REPORTES[sel]}</div>",
                unsafe_allow_html=True)

    # Opciones específicas por reporte
    fname = ""
    data_fn = None

    if sel == "📊 Resumen General":
        fname = f"ResumenGeneral_{hoy.strftime('%Y%m%d')}.xlsx"
        data_fn = reporte_resumen

    elif sel == "📅 Cuotas del Mes":
        MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                 "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        col1, col2, col3 = st.columns([1.5, 1, 2])
        with col1:
            mes_sel = st.selectbox("Mes", range(1,13), index=hoy.month-1,
                                   format_func=lambda m: MESES[m-1], key="mes_r")
        with col2:
            anio_sel = st.number_input("Año", min_value=2020, max_value=2035,
                                       value=hoy.year, step=1, key="anio_r")
        with col3:
            inc_atr = st.checkbox("Incluir cuotas atrasadas acumuladas", value=True, key="inc_atr")
        fname = f"CuotasMes_{mes_sel:02d}-{int(anio_sel)}.xlsx"
        data_fn = lambda: reporte_cuotas_mes(int(anio_sel), mes_sel, inc_atr)

    elif sel == "💳 Historial de Pagos":
        fname = f"HistorialPagos_{hoy.strftime('%Y%m%d')}.xlsx"
        data_fn = reporte_historial

    elif sel == "⚠️ Clientes en Mora":
        fname = f"ClientesMora_{hoy.strftime('%Y%m%d')}.xlsx"
        data_fn = reporte_mora

    elif sel == "📋 Plan de Pagos por Cliente":
        fname = f"PlanPorCliente_{hoy.strftime('%Y%m%d')}.xlsx"
        data_fn = reporte_plan_por_cliente

    elif sel == "🔖 Clientes en Reserva":
        fname = f"Reservas_{hoy.strftime('%Y%m%d')}.xlsx"
        data_fn = reporte_reservas

    if data_fn:
        st.download_button(
            label="⬇️ Descargar Excel",
            data=data_fn(),
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_reporte",
            type="primary",
        )
